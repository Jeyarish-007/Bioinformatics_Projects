"""
# Clinical follow-up mapping – Robust Web Interface
- Always includes all patients in output, even with invalid dates.
- Shows list of cases with date issues for manual review.
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Alignment, Protection
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil.relativedelta import relativedelta
import math
from copy import copy
import io

# ---------------------------
#  Configuration
# ---------------------------
MONTH_INTERVAL = 2
START_MONTH = 1

# ---------------------------
#  Helper functions
# ---------------------------
def find_header_row(sheet, search_term='Lab Id'):
    for idx, row in enumerate(sheet.iter_rows(values_only=False)):
        if any(cell.value and search_term.lower() in str(cell.value).lower() for cell in row):
            return idx
    raise ValueError(f"Header row containing '{search_term}' not found.")

def get_column_indices(header_cells):
    headers = [str(cell.value).strip() if cell.value else '' for cell in header_cells]
    
    baseline_date_idx = None
    for i, h in enumerate(headers):
        if h.lower() == 'date':
            next1 = headers[i+1] if i+1 < len(headers) else ''
            next2 = headers[i+2] if i+2 < len(headers) else ''
            if next1.lower() == 'case no.' and next2.lower() == 'name':
                baseline_date_idx = i
                break
    if baseline_date_idx is None:
        raise ValueError("Could not locate baseline 'Date' column.")
    
    case_no_idx = baseline_date_idx + 1
    name_idx = baseline_date_idx + 2
    age_idx = baseline_date_idx + 3
    sex_idx = baseline_date_idx + 4
    baseline_idx = baseline_date_idx + 5
    nature_idx = baseline_date_idx + 6
    separated_by_idx = baseline_date_idx + 7
    
    lab_id_idx = next((i for i, h in enumerate(headers) if h.lower() == 'lab id'), 0)
    
    followup_date_indices = [
        i for i, h in enumerate(headers)
        if h.upper() == 'DATE' and i != baseline_date_idx
    ]
    
    return {
        'lab_id': lab_id_idx,
        'baseline_date': baseline_date_idx,
        'case_no': case_no_idx,
        'name': name_idx,
        'age': age_idx,
        'sex': sex_idx,
        'baseline': baseline_idx,
        'nature': nature_idx,
        'separated_by': separated_by_idx,
        'followup_dates': followup_date_indices,
        'header_cells': header_cells
    }

def safe_parse_date(value):
    """Parse date with robust error handling. Returns None if invalid."""
    if pd.isna(value) or str(value).strip() == '':
        return None
    try:
        dt = pd.to_datetime(value, dayfirst=True, errors='coerce')
        if pd.isna(dt):
            return None
        # Basic sanity: year between 1900 and 2100
        if dt.year < 1900 or dt.year > 2100:
            return None
        return dt
    except:
        return None

def safe_month_diff_ceil(d1, d2):
    """Calculate ceiling months, return None if any error."""
    if d1 is None or d2 is None:
        return None
    try:
        r = relativedelta(d2, d1)
        total_months = r.years * 12 + r.months
        if r.days > 0:
            total_months += 1
        return total_months
    except Exception:
        return None

def generate_month_columns(max_months, interval=2, start=1):
    columns = []
    current_start = start
    while current_start <= max_months:
        end = current_start + interval - 1
        col_name = f"{current_start}-{end}_months_followups"
        columns.append((current_start, end, col_name))
        current_start += interval
    return columns

def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)

def process_uploaded_file(uploaded_file):
    wb_input = openpyxl.load_workbook(uploaded_file, data_only=True)
    sheet_input = wb_input.active
    
    header_idx = find_header_row(sheet_input, 'Lab Id')
    header_cells = list(sheet_input[header_idx + 1])
    indices = get_column_indices(header_cells)
    
    patients = []
    issue_cases = []  # Track cases with any date issue
    all_followup_months = []
    
    for row_idx, row_cells in enumerate(sheet_input.iter_rows(min_row=header_idx+2), start=header_idx+2):
        if all(cell.value is None or str(cell.value).strip() == '' for cell in row_cells):
            continue
        
        lab_id = row_cells[indices['lab_id']].value
        case_no = row_cells[indices['case_no']].value
        
        # Parse baseline date – may be None
        baseline_date_cell = row_cells[indices['baseline_date']]
        baseline_date = safe_parse_date(baseline_date_cell.value)
        baseline_valid = baseline_date is not None
        
        if not baseline_valid:
            issue_cases.append((lab_id, case_no, "Invalid/missing baseline date"))
        
        # Collect follow-up dates, tracking any invalid ones
        followup_dates = []
        for idx in indices['followup_dates']:
            val = row_cells[idx].value
            dt = safe_parse_date(val)
            if dt is not None:
                if baseline_valid and dt < baseline_date:
                    # Follow-up before baseline? Still include but flag
                    issue_cases.append((lab_id, case_no, f"Follow-up date before baseline: {val}"))
                followup_dates.append(dt)
            elif val is not None and str(val).strip() != '':
                # Non-empty but invalid date
                issue_cases.append((lab_id, case_no, f"Invalid follow-up date: {val}"))
        
        # Calculate months only if baseline is valid
        months_list = []
        if baseline_valid:
            for fu in followup_dates:
                months = safe_month_diff_ceil(baseline_date, fu)
                if months is not None:
                    months_list.append(months)
                else:
                    issue_cases.append((lab_id, case_no, f"Month calculation error for date: {fu}"))
        
        all_followup_months.extend(months_list)
        
        # Always include patient in output
        patients.append({
            'lab_id': lab_id,
            'baseline_date': baseline_date,  # could be None
            'case_no': case_no,
            'name': row_cells[indices['name']].value,
            'age': row_cells[indices['age']].value,
            'sex': row_cells[indices['sex']].value,
            'baseline': row_cells[indices['baseline']].value,
            'nature': row_cells[indices['nature']].value,
            'separated_by': row_cells[indices['separated_by']].value,
            'followup_dates': followup_dates,
            'style_cells': {
                'lab_id': row_cells[indices['lab_id']],
                'date': baseline_date_cell,
                'case_no': row_cells[indices['case_no']],
                'name': row_cells[indices['name']],
                'age': row_cells[indices['age']],
                'sex': row_cells[indices['sex']],
                'baseline': row_cells[indices['baseline']],
                'nature': row_cells[indices['nature']],
                'separated_by': row_cells[indices['separated_by']]
            }
        })
    
    wb_input.close()
    
    max_month = max(all_followup_months) if all_followup_months else 0
    if max_month == 0:
        max_month = MONTH_INTERVAL
    
    month_ranges = generate_month_columns(max_month, MONTH_INTERVAL, START_MONTH)
    
    output_rows = []
    for patient in patients:
        # Format baseline date string (empty if None)
        date_str = patient['baseline_date'].strftime('%d/%m/%Y') if patient['baseline_date'] else ''
        base_row = {
            'Lab Id': patient['lab_id'],
            'Date': date_str,
            'CASE NO.': patient['case_no'],
            'NAME': patient['name'],
            'AGE': patient['age'],
            'SEX': patient['sex'],
            '(BASELINE)': patient['baseline'],
            'Nature of Sample': patient['nature'],
            'SEPARATED BY': patient['separated_by'],
            'Lab Id_dup': patient['lab_id']
        }
        
        range_dict = {col_name: None for _, _, col_name in month_ranges}
        
        # Only process follow-ups if baseline is valid
        if patient['baseline_date'] is not None:
            for fu_date in patient['followup_dates']:
                months = safe_month_diff_ceil(patient['baseline_date'], fu_date)
                if months is None:
                    continue
                for start, end, col_name in month_ranges:
                    if start <= months <= end:
                        if range_dict[col_name] is None or fu_date > range_dict[col_name]:
                            range_dict[col_name] = fu_date
                        break
        
        for col_name, dt in range_dict.items():
            base_row[col_name] = dt.strftime('%d/%m/%Y') if dt else ''
        
        output_rows.append(base_row)
    
    static_cols = ['Lab Id', 'Date', 'CASE NO.', 'NAME', 'AGE', 'SEX', '(BASELINE)',
                   'Nature of Sample', 'SEPARATED BY', 'Lab Id_dup']
    dynamic_cols = [col for _, _, col in month_ranges]
    columns_order = static_cols + dynamic_cols
    
    df_out = pd.DataFrame(output_rows, columns=columns_order)
    
    output_bytes = io.BytesIO()
    with pd.ExcelWriter(output_bytes, engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name='Sheet1', index=False)
        
        wb_out = writer.book
        ws_out = writer.sheets['Sheet1']
        
        col_map = {col_name: idx+1 for idx, col_name in enumerate(columns_order)}
        
        header_style_map = {
            'Lab Id': indices['lab_id'],
            'Date': indices['baseline_date'],
            'CASE NO.': indices['case_no'],
            'NAME': indices['name'],
            'AGE': indices['age'],
            'SEX': indices['sex'],
            '(BASELINE)': indices['baseline'],
            'Nature of Sample': indices['nature'],
            'SEPARATED BY': indices['separated_by'],
            'Lab Id_dup': indices['lab_id']
        }
        for out_col, in_idx in header_style_map.items():
            source_cell = header_cells[in_idx]
            target_cell = ws_out.cell(row=1, column=col_map[out_col])
            copy_cell_style(source_cell, target_cell)
        
        for i, patient in enumerate(patients, start=2):
            style_cells = patient['style_cells']
            copy_cell_style(style_cells['lab_id'], ws_out.cell(row=i, column=col_map['Lab Id']))
            copy_cell_style(style_cells['date'], ws_out.cell(row=i, column=col_map['Date']))
            copy_cell_style(style_cells['case_no'], ws_out.cell(row=i, column=col_map['CASE NO.']))
            copy_cell_style(style_cells['name'], ws_out.cell(row=i, column=col_map['NAME']))
            copy_cell_style(style_cells['age'], ws_out.cell(row=i, column=col_map['AGE']))
            copy_cell_style(style_cells['sex'], ws_out.cell(row=i, column=col_map['SEX']))
            copy_cell_style(style_cells['baseline'], ws_out.cell(row=i, column=col_map['(BASELINE)']))
            copy_cell_style(style_cells['nature'], ws_out.cell(row=i, column=col_map['Nature of Sample']))
            copy_cell_style(style_cells['separated_by'], ws_out.cell(row=i, column=col_map['SEPARATED BY']))
            copy_cell_style(style_cells['lab_id'], ws_out.cell(row=i, column=col_map['Lab Id_dup']))
        
        for col in ws_out.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_out.column_dimensions[col_letter].width = adjusted_width
    
    output_bytes.seek(0)
    return output_bytes, dynamic_cols, issue_cases

# ---------------------------
#  Streamlit UI
# ---------------------------
st.set_page_config(page_title="Clinical Follow-Up Mapper", page_icon="📊")
st.title("📋 Clinical Follow-Up Date Mapper")

st.markdown("""
### Instructions for Use
1. **Prepare your Excel file** – Must have columns: Lab Id, Date (baseline), CASE NO., NAME, etc.
2. **Click "Browse files"** and select your Excel file.
3. **Wait for processing** – all rows will be included in the output, even if dates are invalid.
4. **Download the processed file**.

Cases with date issues will be listed below for your review; their follow‑up columns will be empty in the output.
""")

uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])

if uploaded_file is not None:
    with st.spinner("Processing file... Please wait."):
        try:
            output_bytes, dynamic_cols, issue_cases = process_uploaded_file(uploaded_file)
            st.success("✅ Processing complete!")
            st.write(f"Generated columns: {', '.join(dynamic_cols)}")
            
            if issue_cases:
                st.warning(f"⚠️ {len(issue_cases)} issue(s) detected (output rows included but follow‑up columns empty):")
                issue_df = pd.DataFrame(issue_cases, columns=["Lab Id", "CASE NO.", "Issue"])
                st.dataframe(issue_df, use_container_width=True)
            else:
                st.info("All cases processed without date issues.")
            
            st.download_button(
                label="📥 Download Processed File",
                data=output_bytes,
                file_name="followup_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"❌ Error: {str(e)}")
            st.exception(e)

# ---------------------------
#  Footer
# ---------------------------
st.markdown("---")
st.markdown("Developed by Jeyarish")